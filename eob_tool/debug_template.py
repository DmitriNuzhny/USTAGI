from __future__ import annotations

import argparse
from openpyxl import load_workbook


LABELS = [
    "Property Address:",
    "Building Use:",
    "Date Placed in Service:",
    "Cost Basis:",
    "Land Allocation:",
    "Building Basis:",
    "Improvements included in the Study:",
    "Basis for Cost Segregation:",
    "Total Accelerated:",
    "Tax Savings Benefit at a 40% Income Tax Rate:",
    "Estimated Additional Depreciation Expense:",
]

def norm(x):
    return str(x).strip().lower() if x is not None else ""

def find_label_and_value_cell(ws, label: str):
    target = norm(label)
    for row in ws.iter_rows():
        for cell in row:
            if norm(cell.value) == target:
                r, c = cell.row, cell.column
                # scan right for the "visible value cell" (first non-empty)
                for cc in range(c + 1, ws.max_column + 1):
                    v = ws.cell(r, cc).value
                    if v is None or norm(v) == "":
                        continue
                    return (cell.coordinate, ws.cell(r, cc).coordinate, ws.cell(r, cc).value)
                # fallback
                return (cell.coordinate, ws.cell(r, c + 1).coordinate, ws.cell(r, c + 1).value)
    return None

def find_tax_year_table(ws):
    # find the header row containing "Tax Year"
    header_row = None
    year_col = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(r, c).value) == "tax year":
                header_row, year_col = r, c
                break
        if header_row:
            break
    if not header_row:
        return None

    # identify columns in that header row
    def find_col_contains(token: str):
        t = token.lower()
        for c in range(1, ws.max_column + 1):
            v = norm(ws.cell(header_row, c).value)
            if t in v and v != "":
                return c
        return None

    cols = {
        "header_row": header_row,
        "year_col": year_col,
        "col_5yr": find_col_contains("5 years"),
        "col_7yr": find_col_contains("7 years"),
        "col_15yr": find_col_contains("15 years"),
        "col_275": find_col_contains("27.5"),
        "col_39": find_col_contains("39"),
        "col_with_css": find_col_contains("depr") ,  # may hit first "Depr."
        "col_without_css": None,
    }

    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(header_row, c).value)
        if "without" in v and "depr" in v:
            cols["col_without_css"] = c
            break

    return cols

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()

    wb = load_workbook(args.template)
    ws = wb[args.sheet] if args.sheet else wb.active

    print(f"Template: {args.template}")
    print(f"Sheet: {ws.title}")
    print("\n[Label -> Value cell guesses]")
    for lab in LABELS:
        res = find_label_and_value_cell(ws, lab)
        if res:
            lab_cell, val_cell, current = res
            print(f"{lab}  label={lab_cell}  value={val_cell}  current={current!r}")
        else:
            print(f"{lab}  NOT FOUND")

    print("\n[Tax year table detection]")
    tbl = find_tax_year_table(ws)
    if not tbl:
        print("Tax Year header NOT FOUND")
    else:
        print(tbl)

if __name__ == "__main__":
    main()
