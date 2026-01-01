from __future__ import annotations

from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D


RES_TEMPLATE = Path("templates") / "estimator_residential.xlsx"
COM_TEMPLATE = Path("templates") / "estimator_commercial.xlsx"


# These come directly from your debug_template output.
RES_MAP = {
    "property_address": "G9",
    "building_use": "G10",
    "date_in_service": "F11",
    "cost_basis": "G12",
    # Land Allocation was weird (value cell guess D13 is a formula). We'll support both:
    "land_allocation_text": "D13",   # e.g., "Per Depreciation Schedule"
    "land_allocation_amount": "G13", # if you want the $ value
    "building_basis": "G14",
    "improvements_included": "G15",
    "basis_for_cost_seg": "G16",
    "total_accelerated": "G19",
    "tax_savings_total": "G20",
    "estimated_addl_depr": "G22",
    # The “tax savings” for the additional depreciation is usually the next line.
    # If it’s not, adjust this to match the template.
    "tax_savings_addl": "G23",    "tier": "G21",
    "table": {
        "start_row": 27,
        "n_years": 29,          # 2021–2049 (adjust if your template has more/less)
        "col_year": 2,          # B
        "col_5yr": 3,           # C
        "col_7yr": 4,           # D
        "col_15yr": 5,          # E
        "col_long": 6,          # F
        "col_with_css": 7,      # G
        "col_without_css": 8,   # H
    },
}

COM_MAP = {
    "property_address": "G9",
    "building_use": "G10",
    "date_in_service": "F11",
    "cost_basis": "G12",
    "land_allocation_text": "D13",
    "land_allocation_amount": "G13",
    "building_basis": "G14",
    "improvements_included": "G15",
    "basis_for_cost_seg": "G16",
    "total_accelerated": "G19",
    "tax_savings_total": "G20",
    "estimated_addl_depr": "G22",
    "tax_savings_addl": "G23",

    "table": {
        "start_row": 27,
        "n_years": 29,          # 2021–2049 (adjust if your template has more/less)
        "col_year": 2,          # B
        "col_5yr": 3,           # C
        "col_7yr": 4,           # D
        "col_15yr": 5,          # E
        "col_long": 6,          # F
        "col_with_css": 7,      # G
        "col_without_css": 8,   # H
    },
}


EMU_PER_PIXEL = 9525

def ensure_logo_exact(ws, logo_path: str = "templates/ustagi_logo.png") -> None:
    """
    Insert logo with exact placement measured from the template.
    Template anchor readout:
      TwoCellAnchor
      from: col=3,row=0,colOff=647700,rowOff=0
      size: 600x264 px
    """
    p = Path(logo_path)
    if not p.exists():
        return

    # Prevent duplicates
    try:
        ws._images = []
    except Exception:
        pass

    img = Image(str(p))

    # --- Exact placement from your template ---
    FROM_COL = 3               # 0-based -> D
    FROM_ROW = 0               # 0-based -> row 1
    COL_OFF_EMU = 647700       # = 68 px
    ROW_OFF_EMU = 0
    WIDTH_PX = 262
    HEIGHT_PX = 115
    # -----------------------------------------

    marker = AnchorMarker(
        col=FROM_COL,
        colOff=COL_OFF_EMU,
        row=FROM_ROW,
        rowOff=ROW_OFF_EMU,
    )

    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            cx=WIDTH_PX * EMU_PER_PIXEL,
            cy=HEIGHT_PX * EMU_PER_PIXEL,
        ),
    )

    ws.add_image(img)


def _as_dict(obj: Any) -> Dict[str, Any]:
    if isinstance(obj, dict):
        return obj
    if is_dataclass(obj):
        return asdict(obj)
    if hasattr(obj, "__dict__"):
        return dict(obj.__dict__)
    raise TypeError(f"Unsupported result type: {type(obj)}")


def _ws_set(ws: Worksheet, addr: str, value: Any) -> None:
    # Overwrite formulas/#REF! with a concrete value.
    ws[addr].value = value


def _fill_table(ws: Worksheet, cfg: Dict[str, Any], yearly: Dict[int, Dict[str, Any]]) -> None:
    start_row = int(cfg["start_row"])      # 27
    n_years = int(cfg["n_years"])          # how many rows to fill (e.g., 29 for 2021-2049)
    start_year = int(cfg["start_year"])    # e.g., YEAR(date placed in service)

    col_year = int(cfg["col_year"])        # B=2
    c5 = int(cfg["col_5yr"])               # C=3
    c7 = int(cfg["col_7yr"])               # D=4
    c15 = int(cfg["col_15yr"])             # E=5
    clong = int(cfg["col_long"])           # F=6
    cwith = int(cfg["col_with_css"])       # G=7
    cwithout = int(cfg["col_without_css"]) # H=8

    # Initialize sums
    sum_5yr = 0.0
    sum_7yr = 0.0
    sum_15yr = 0.0
    sum_long = 0.0
    sum_with = 0.0
    sum_without = 0.0

    # Write literal years + values row-by-row
    for i in range(n_years):
        year = start_year + i
        r = start_row + i

        ws.cell(r, col_year).value = year

        row = yearly.get(year, {})
        val_5yr = row.get("5yr", 0.0) or 0.0
        val_7yr = row.get("7yr", 0.0) or 0.0
        val_15yr = row.get("15yr", 0.0) or 0.0
        val_long = row.get("long", 0.0) or 0.0
        val_with = row.get("with_css", 0.0) or 0.0
        val_without = row.get("without_css", 0.0) or 0.0

        ws.cell(r, c5).value = val_5yr
        ws.cell(r, c7).value = val_7yr
        ws.cell(r, c15).value = val_15yr
        ws.cell(r, clong).value = val_long
        ws.cell(r, cwith).value = val_with
        ws.cell(r, cwithout).value = val_without

        sum_5yr += val_5yr
        sum_7yr += val_7yr
        sum_15yr += val_15yr
        sum_long += val_long
        sum_with += val_with
        sum_without += val_without

    # Add totals row
    totals_row = start_row + n_years
    ws.cell(totals_row, col_year).value = "Total"
    ws.cell(totals_row, c5).value = sum_5yr
    ws.cell(totals_row, c7).value = sum_7yr
    ws.cell(totals_row, c15).value = sum_15yr
    ws.cell(totals_row, clong).value = sum_long
    ws.cell(totals_row, cwith).value = sum_with
    ws.cell(totals_row, cwithout).value = sum_without



def _require_payload_shape(res: Dict[str, Any], mode: str) -> None:
    """
    Make it impossible to silently output a blank template.
    We require:
      res["summary"] : dict
      res["yearly"]  : dict[int -> dict]
    """
    if "summary" not in res or "yearly" not in res:
        keys = sorted(res.keys())
        raise ValueError(
            f"{mode} compute result does not include required keys 'summary' and 'yearly'.\n"
            f"Top-level keys present: {keys}\n\n"
            "Fix: add an adapter at the end of compute_residential/compute_commercial to return:\n"
            "  {'summary': {...}, 'yearly': {year: {'5yr':..,'7yr':..,'15yr':..,'long':..,'with_css':..,'without_css':..}}}\n"
        )


def _fill_estimator(ws: Worksheet, mapping: Dict[str, Any], result_obj: Any, mode: str) -> None:
    # If sheet is protected without a password, this allows writing.
    try:
        ws.protection.sheet = False
    except Exception:
        pass

    res = _as_dict(result_obj)
    _require_payload_shape(res, mode)

    summary: Dict[str, Any] = res["summary"] or {}
    yearly: Dict[int, Dict[str, Any]] = res["yearly"] or {}

    _ws_set(ws, mapping["property_address"], summary.get("property_address"))
    _ws_set(ws, mapping["building_use"], summary.get("building_use"))
    _ws_set(ws, mapping["date_in_service"], summary.get("date_placed_in_service"))
    _ws_set(ws, mapping["cost_basis"], summary.get("cost_basis"))

    # Land allocation: allow either text, amount, or both
    if "land_allocation_text" in mapping and summary.get("land_allocation_text") is not None:
        _ws_set(ws, mapping["land_allocation_text"], summary.get("land_allocation_text"))
    if "land_allocation_amount" in mapping:
        _ws_set(ws, mapping["land_allocation_amount"], summary.get("land_allocation_amount", ""))

    _ws_set(ws, mapping["building_basis"], summary.get("building_basis"))
    _ws_set(ws, mapping["improvements_included"], summary.get("improvements_included"))
    _ws_set(ws, mapping["basis_for_cost_seg"], summary.get("basis_for_cost_segregation"))

    _ws_set(ws, mapping["total_accelerated"], summary.get("total_accelerated"))
    _ws_set(ws, mapping["tax_savings_total"], summary.get("tax_savings_40pct_total_accel"))
    _ws_set(ws, mapping["estimated_addl_depr"], summary.get("estimated_additional_depr"))
    _ws_set(ws, mapping["tax_savings_addl"], summary.get("tax_savings_40pct_addl_depr"))

    if "tier" in mapping:
        _ws_set(ws, mapping["tier"], summary.get("tier"))

    # Set start_year for table
    dps = summary.get("date_placed_in_service")
    dps = "" if dps is None else str(dps).strip()

    if dps.lower() in ("none", "null", "nan", ""):
        dps = ""

    if len(dps) >= 4 and dps[:4].isdigit():
        start_year = int(dps[:4])
    else:
        start_year = min(yearly.keys()) if yearly else 2021

    # Set n_years to fill up to 2050
    last_year = 2050
    n_years = last_year - start_year + 1
    mapping["table"]["start_year"] = start_year
    mapping["table"]["n_years"] = n_years

    _fill_table(ws, mapping["table"], yearly or {})


def _clear_excel_errors(wb) -> None:
    """
    Clears cells that contain broken formulas like #REF! (and optionally #VALUE!),
    so the delivered workbook doesn't show errors when we're not populating those sections.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and ("#REF!" in v or "#VALUE!" in v):
                    cell.value = None


def write_residential_workbook(result: Any, out_path: Path | str) -> None:
    out_path = Path(out_path)
    wb = load_workbook(RES_TEMPLATE)
    ws = wb.active  # single sheet

    _fill_estimator(ws, RES_MAP, result, mode="residential")

    _clear_excel_errors(wb)   # <-- add this

    ensure_logo_exact(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_commercial_workbook(result: Any, out_path: Path | str) -> None:
    out_path = Path(out_path)
    wb = load_workbook(COM_TEMPLATE)
    ws = wb.active  # single sheet

    _fill_estimator(ws, COM_MAP, result, mode="commercial")

    _clear_excel_errors(wb)   # <-- add this

    ensure_logo_exact(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
