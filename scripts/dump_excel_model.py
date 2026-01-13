from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

def is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")

def dump_sheet_formulas(wb_path: str, sheet_name: str, out_txt: str) -> None:
    wb = load_workbook(wb_path, data_only=False)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows():
        for cell in row:
            if is_formula(cell.value):
                rows.append(f"{sheet_name}!{cell.coordinate} = {cell.value}")

    Path(out_txt).write_text("\n".join(rows), encoding="utf-8")
    print(f"Wrote {len(rows)} formulas to {out_txt}")

def dump_sheet_constants(wb_path: str, sheet_name: str, out_txt: str) -> None:
    wb = load_workbook(wb_path, data_only=False)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if not is_formula(v):
                rows.append(f"{sheet_name}!{cell.coordinate} = {v!r}")

    Path(out_txt).write_text("\n".join(rows), encoding="utf-8")
    print(f"Wrote {len(rows)} constants to {out_txt}")

if __name__ == "__main__":
    # Adjust these:
    wb_path = "templates/Lookback Template.xlsx"  # put your real file path here
    dump_sheet_formulas(wb_path, "27.5 Estimate", "outputs/formulas_estimate.txt")

    # Also dump likely input sheets once you know their names, e.g. "Inputs"
    # dump_sheet_constants(wb_path, "Inputs", "outputs/constants_inputs.txt")
