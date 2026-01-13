#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from openpyxl import load_workbook

def read_range(ws, start_cell: str, end_cell: str):
    cells = ws[start_cell:end_cell]
    out = []
    for row in cells:
        out_row = []
        for c in row:
            out_row.append(c.value)
        out.append(out_row)
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--wb", required=True, help="Path to Excel file that has been calculated by Excel")
    ap.add_argument("--sheet", required=True, help="e.g. 'Client Summary' or '27.5 Estimate'")
    ap.add_argument("--range", required=True, help="e.g. C27:H55")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.wb, data_only=True)
    ws = wb[args.sheet]
    start, end = args.range.split(":")
    data = read_range(ws, start, end)

    payload = {
        "workbook": args.wb,
        "sheet": args.sheet,
        "range": args.range,
        "values": data,
    }

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)

    print(f"Wrote extracted values to: {args.out}")

if __name__ == "__main__":
    main()