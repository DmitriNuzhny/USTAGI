#!/usr/bin/env python3
"""
Local Monday.com development CLI for USTAGI Lookback Template.

Usage:
  python -m scripts.lookback_local --list-items
  python -m scripts.lookback_local --item-id 123 [--out outputs/file.xlsx]
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import load_workbook

# Import from app.py (production webhook code)
from app import (
    _env,
    _monday_graphql,
    fetch_board_schema,
    build_colid_to_title,
    fetch_item_column_values,
)


def fetch_item_name(token: str, item_id: int) -> str:
    query = """
    query ($item_id: [ID!]!) {
      items(ids: $item_id) {
        name
      }
    }
    """
    data = _monday_graphql(token, query, {"item_id": [int(item_id)]})
    items = data.get("items") or []
    if not items:
        raise RuntimeError(f"No item returned for item_id={item_id}")
    return items[0].get("name") or ""


def list_board_items(token: str, board_id: int) -> list[dict]:
    query = """
    query ($board_id: [ID!]!) {
      boards(ids: $board_id) {
        items_page(limit: 100) {
          items { id name }
        }
      }
    }
    """
    data = _monday_graphql(token, query, {"board_id": [int(board_id)]})
    boards = data.get("boards") or []
    if not boards:
        raise RuntimeError(f"No board returned for board_id={board_id}")
    items = boards[0].get("items_page", {}).get("items", []) or []
    return [{"id": int(it["id"]), "name": it.get("name") or ""} for it in items]


def lookback_item_to_fields(column_values: list[dict], colid_to_title: dict[str, str]) -> dict:
    """
    Extract fields for lookback template from Monday item.
    """
    out: dict = {}

    by_title: dict[str, dict] = {}
    for cv in column_values:
        col_id = cv.get("id")
        title = colid_to_title.get(col_id, "")
        by_title[title.strip().lower()] = cv

    def take(title: str) -> str | None:
        cv = by_title.get(title.lower())
        if not cv:
            return None
        txt = (cv.get("text") or "").strip()
        return txt if txt else None

    def take_float(title: str) -> float | None:
        txt = take(title)
        if txt:
            try:
                return float(txt.replace(",", "").replace("$", ""))
            except ValueError:
                pass
        return None

    out["property_address"] = take("Property Address")
    out["date_placed_in_service"] = take("Date Placed in Service") or take("In Service Date")
    out["building_basis"] = take_float("Building Basis")
    out["land_allocation"] = take_float("Land Allocation")
    out["imp_before"] = take_float("Imp. made BEFORE ISD")
    out["imp_after"] = take_float("Imp. made AFTER ISD")
    out["accumulated_depreciation"] = take_float("Accumulated Depreciation")
    out["property_type"] = take("Property Type") or take("Occupancy Type")
    out["building_use"] = take("Building Use")

    return out


def normalize_lookback_fields(fields: dict) -> dict:
    """
    Normalize fields for lookback template.
    """
    norm = {}

    # study_tax_year = year(date_placed_in_service)
    dps = fields.get("date_placed_in_service")
    if dps:
        try:
            dt = datetime.fromisoformat(dps)
            norm["study_tax_year"] = dt.year
        except ValueError:
            pass

    # improvements = before + after (treat missing as 0)
    imp_before = fields.get("imp_before") or 0
    imp_after = fields.get("imp_after") or 0
    norm["improvements"] = imp_before + imp_after

    # cost_basis_for_template = building_basis + land_allocation
    bb = fields.get("building_basis") or 0
    la = fields.get("land_allocation") or 0
    norm["cost_basis_for_template"] = bb + la

    # building_use_map: if building_use is "SFR" set "Single Family Residence" else pass-through
    bu = fields.get("building_use")
    if bu == "SFR":
        norm["building_use"] = "Single Family Residence"
    else:
        norm["building_use"] = bu

    # Property type: Residential or Commercial
    pt = fields.get("property_type", "").lower()
    if "res" in pt:
        norm["property_type"] = "Residential"
    elif "com" in pt:
        norm["property_type"] = "Commercial"
    else:
        norm["property_type"] = fields.get("property_type")

    # Pass through others
    norm["property_address"] = fields.get("property_address")
    norm["date_placed_in_service"] = fields.get("date_placed_in_service")
    norm["land_allocation"] = fields.get("land_allocation")
    norm["accumulated_depreciation"] = fields.get("accumulated_depreciation")

    return norm


def main() -> None:
    load_dotenv()

    ap = argparse.ArgumentParser(prog="lookback_local", description="Local Monday.com development CLI for Lookback Template")
    ap.add_argument("--list-items", action="store_true", help="List items on the test board")
    ap.add_argument("--item-id", type=int, help="Item ID to process locally")
    ap.add_argument("--out", help="Output path (must be under outputs/). If omitted, auto-generates.")

    args = ap.parse_args()

    if not args.list_items and not args.item_id:
        ap.error("Must specify --list-items or --item-id")

    token = _env("MONDAY_API_TOKEN")
    board_id = int(_env("MONDAY_BOARD_ID"))

    if args.list_items:
        items = list_board_items(token, board_id)
        for it in items:
            print(f"{it['name']} : {it['id']}")
        return

    # --item-id mode
    item_id = args.item_id
    item_name = fetch_item_name(token, item_id)
    if not item_name:
        raise RuntimeError(f"Item {item_id} has no name")

    safe_name = item_name.replace(" ", "_").replace("/", "_")

    # Fetch fields
    board_schema = fetch_board_schema(token, board_id)
    colid_to_title = build_colid_to_title(board_schema)
    col_vals = fetch_item_column_values(token, item_id)
    fields = lookback_item_to_fields(col_vals, colid_to_title)
    norm = normalize_lookback_fields(fields)

    # Load template
    template_path = Path("templates") / "Lookback Template.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb["Client Summary"]

    # Write to cells
    ws["G12"] = norm.get("property_address")
    dps = norm.get("date_placed_in_service")
    if dps:
        try:
            ws["G13"] = datetime.fromisoformat(dps).date()
        except ValueError:
            ws["G13"] = dps
    ws["G14"] = norm.get("cost_basis_for_template")
    ws["G15"] = norm.get("land_allocation")
    ws["G17"] = norm.get("improvements")
    ws["G19"] = norm.get("accumulated_depreciation")
    ws["G20"] = norm.get("property_type")
    ws["G21"] = norm.get("building_use")
    ws["N13"] = norm.get("study_tax_year")

    # Ensure recalc
    wb.calculation.fullCalcOnLoad = True

    if args.out:
        out_path = Path(args.out)
        if not out_path.is_relative_to(Path("outputs")):
            raise RuntimeError("Output path must be under outputs/")
    else:
        # Auto-generate
        out_path = Path("outputs") / f"{safe_name}__{item_id}__lookback.xlsx"

    Path("outputs").mkdir(exist_ok=True)
    wb.save(out_path)
    print(f"Generated {out_path}")


if __name__ == "__main__":
    main()