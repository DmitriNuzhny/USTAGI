#!/usr/bin/env python3

from openpyxl import load_workbook
from dataclasses import dataclass
from typing import Any
from pathlib import Path

@dataclass
class Hit:
    kind: str           # "FORMULA" or "VALUE"
    sheet: str
    cell: str
    text: str

KEYWORDS = ["tier", "SFR", "SFR$", "SFR$$", "SFR$$$", "Client Summary", "Study Data"]

def _is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")

def _contains_any(text: str, needles: list[str]) -> bool:
    t = text.lower()
    return any(n.lower() in t for n in needles)

def inspect_workbook(wb_path: str) -> dict:
    wb = load_workbook(wb_path, data_only=False)

    hits: list[Hit] = []
    formulas: list[Hit] = []
    sheets = wb.sheetnames
    named_ranges = [f"{name} = {defn.attr_text}" for name, defn in wb.defined_names.items()]

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                coord = f"{ws.title}!{cell.coordinate}"
                if _is_formula(val):
                    formulas.append(Hit("FORMULA", ws.title, cell.coordinate, val))
                if _contains_any(val, KEYWORDS):
                    kind = "FORMULA" if _is_formula(val) else "VALUE"
                    hits.append(Hit(kind, ws.title, cell.coordinate, val))

    return {
        "path": wb_path,
        "sheets": sheets,
        "named_ranges": named_ranges,
        "formulas": formulas,
        "keyword_hits": hits,
    }

def print_report(report: dict):
    print("=" * 120)
    print(f"WORKBOOK: {report['path']}")
    print("=" * 120)
    print(f"Sheets ({len(report['sheets'])}):")
    for s in report["sheets"]:
        print(f"  - {s}")

    print("-" * 40)
    print(f"Named Ranges ({len(report['named_ranges'])}):")
    for r in report["named_ranges"]:
        print(f"  - {r}")

    print("-" * 40)
    print(f"Keyword Matches ({len(report['keyword_hits'])}):")
    for hit in report["keyword_hits"]:
        print(f"{hit.kind:7}  {hit.sheet}!{hit.cell}  {hit.text}")

    print("-" * 40)
    print(f"Total formulas: {len(report['formulas'])}")
    print("=" * 120)
    print()

def compare_keyword_presence(rep1: dict, rep2: dict, keywords: list[str]):
    print("=" * 120)
    print("Keyword Presence Comparison")
    print("=" * 120)
    for kw in keywords:
        f1 = sum(1 for h in rep1["keyword_hits"] if kw.lower() in h.text.lower())
        f2 = sum(1 for h in rep2["keyword_hits"] if kw.lower() in h.text.lower())
        print(f"- {kw!r}")
        print(f"  {Path(rep1['path']).name}: {f1}")
        print(f"  {Path(rep2['path']).name}: {f2}")
    print("=" * 120)
    print()

def main():
    lookback_file = "templates/Lookback Template.xlsx"
    default_file = "inputs/commercial_estimator_default_settings.xlsx"

    lookback = inspect_workbook(lookback_file)
    default = inspect_workbook(default_file)

    print_report(lookback)
    print_report(default)
    compare_keyword_presence(lookback, default, KEYWORDS)

if __name__ == "__main__":
    main()

