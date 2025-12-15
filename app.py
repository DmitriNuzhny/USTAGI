import io
import mimetypes
import os
import json
import logging
from datetime import datetime

import requests
from dotenv import load_dotenv
from fastapi import FastAPI, Request, Response
from fastapi.responses import JSONResponse
from openpyxl import Workbook

# -----------------------
# Setup
# -----------------------
load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger("monday-export-eob")

app = FastAPI()


def _env(name: str, required: bool = True) -> str:
    val = os.getenv(name, "").strip()
    if required and not val:
        raise RuntimeError(f"{name} missing/invalid")
    return val


def _safe_json(obj) -> str:
    try:
        return json.dumps(obj, indent=2)
    except Exception:
        return str(obj)


def _find_item_id(payload: dict) -> int | None:
    """
    monday app events payload shapes vary.
    Try a few known paths, then do a recursive search for 'itemId'/'item_id'.
    """
    # Common candidates
    candidates = [
        ("event", "itemId"),
        ("event", "item_id"),
        ("payload", "itemId"),
        ("payload", "item_id"),
        ("data", "itemId"),
        ("data", "item_id"),
    ]

    for path in candidates:
        cur = payload
        ok = True
        for key in path:
            if isinstance(cur, dict) and key in cur:
                cur = cur[key]
            else:
                ok = False
                break
        if ok:
            try:
                return int(cur)
            except Exception:
                pass

    # Recursive search
    def walk(x):
        if isinstance(x, dict):
            for k, v in x.items():
                lk = str(k).lower()
                if lk in ("itemid", "item_id"):
                    yield v
                yield from walk(v)
        elif isinstance(x, list):
            for v in x:
                yield from walk(v)

    for v in walk(payload):
        try:
            return int(v)
        except Exception:
            continue

    return None


def _make_test_excel_bytes(item_id: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "EOB Export"

    ws["A1"] = "Hello from Export EOB"
    ws["A2"] = "Item ID"
    ws["B2"] = item_id
    ws["A3"] = "Generated at"
    ws["B3"] = datetime.now().isoformat(timespec="seconds")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _monday_upload_file_to_column(
    api_token: str,
    item_id: int,
    column_id: str,
    file_bytes: bytes,
    filename: str,
) -> dict:
    """
    monday file upload format for /v2/file:
      - multipart fields:
          query: 'mutation ($file: File!) { add_file_to_column(item_id: X, column_id: "Y", file: $file) { id } }'
          map:  '{"image":"variables.file"}'
          image: <file>
    """
    url = "https://api.monday.com/v2/file"

    # Put item_id + column_id directly in the query; only variable is $file
    query = (
        'mutation ($file: File!) { '
        f'add_file_to_column(item_id: {int(item_id)}, column_id: "{column_id}", file: $file) '
        '{ id } }'
    )

    # monday expects a JSON string mapping the file field name -> variables.file
    map_field = {"image": "variables.file"}

    content_type, _ = mimetypes.guess_type(filename)
    if not content_type:
        content_type = "application/octet-stream"

    headers = {"Authorization": api_token}  # don't set Content-Type manually

    files = {
        "query": (None, query),
        "map": (None, json.dumps(map_field)),
        "image": (filename, file_bytes, content_type),
    }

    resp = requests.post(url, headers=headers, files=files, timeout=60)

    try:
        payload = resp.json()
    except Exception:
        raise RuntimeError(f"monday upload: non-JSON response {resp.status_code}: {resp.text[:500]}")

    if resp.status_code >= 400:
        raise RuntimeError(f"monday upload HTTP {resp.status_code}: {payload}")

    if payload.get("errors"):
        raise RuntimeError(f"monday upload returned errors: {payload}")

    return payload


# -----------------------
# Routes
# -----------------------
@app.get("/health")
def health():
    return {"ok": True}


@app.post("/monday/subscribe")
async def monday_subscribe(request: Request):
    body = await request.json()
    log.info(f"SUBSCRIBE: {_safe_json(body)}")
    # For app events / automations, returning 200 is sufficient.
    return JSONResponse({"ok": True})


@app.post("/monday/unsubscribe")
async def monday_unsubscribe(request: Request):
    body = await request.json()
    log.info(f"UNSUBSCRIBE: {_safe_json(body)}")
    return JSONResponse({"ok": True})


@app.post("/monday/webhook/export-eob")
async def export_eob_webhook(request: Request):
    """
    Called when the automation/button triggers.
    Action: upload a test xlsx to the triggering item, into the file column you specify.
    """
    body = await request.json()
    log.info(f"EXPORT-EOB WEBHOOK: {_safe_json(body)}")

    # Required config
    file_column_id = _env("MONDAY_FILE_COLUMN_ID")  # e.g. "files" or your file column id
    # (Optional) use a stable name for testing
    filename = os.getenv("TEST_EXPORT_FILENAME", "export_eob_test.xlsx").strip() or "export_eob_test.xlsx"

    item_id = _find_item_id(body)
    if not item_id:
        # Permanent misconfig: return 400 to indicate bad request
        log.error("Could not find itemId in webhook payload. No upload performed.")
        return JSONResponse(status_code=400, content={"error": "missing_item_id"})

    # Create a test Excel and upload
    try:
        excel_bytes = _make_test_excel_bytes(item_id=item_id)
        upload_result = _monday_upload_file_to_column(
            api_token=_env("MONDAY_API_TOKEN"),
            item_id=item_id,
            column_id=file_column_id,
            file_bytes=excel_bytes,
            filename=filename,
        )
        log.info(f"Uploaded file to item {item_id} column {file_column_id}: {_safe_json(upload_result)}")
        return JSONResponse({"ok": True, "uploaded": True, "itemId": item_id})
    except Exception as e:
        # Return 200 to avoid aggressive retries while you're iterating (you can change to 500 later if you want).
        log.exception(f"Upload failed: {e}")
        return JSONResponse({"ok": True, "uploaded": False, "itemId": item_id, "error": str(e)})


if __name__ == "__main__":
    import uvicorn

    # run: python app.py
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
